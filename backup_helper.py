import streamlit as st
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

def create_backup_page():
    import streamlit as st
    """Create a backup/export page for database"""
    st.header("🔄 Database Backup & Export")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📊 Current Database Stats")
        conn = sqlite3.connect('chords_crm.db')
        
        # Get counts
        students_count = pd.read_sql_query("SELECT COUNT(*) as count FROM students", conn).iloc[0]['count']
        payments_count = pd.read_sql_query("SELECT COUNT(*) as count FROM payments", conn).iloc[0]['count']
        total_revenue = pd.read_sql_query("SELECT COALESCE(SUM(amount), 0) as total FROM payments", conn).iloc[0]['total']
        
        st.metric("Active Students", students_count)
        st.metric("Total Payments", payments_count)
        st.metric("Total Revenue", f"₹{total_revenue:,.0f}")
        
        conn.close()
    
    with col2:
        st.subheader("💾 Export Options")
        
        if st.button("📥 Download Students CSV"):
            conn = sqlite3.connect('chords_crm.db')
            
            # Based on your latest data: Aditya Palagummi, 43, 9701802225, [empty], [empty], [empty], aditya.cn@gmail.com, 22/08/82, Keyboard
            # The actual database columns seem to be: full_name, age, mobile, [empty], [empty], [empty], email, date_of_birth, instrument
            df = pd.read_sql_query("""
                SELECT 
                    full_name,
                    age,
                    mobile,
                    COALESCE(email, '') as email,
                    COALESCE(date_of_birth, '') as date_of_birth,
                    COALESCE(sex, '') as sex,
                    COALESCE(instrument, '') as instrument,
                    COALESCE(class_plan, '') as class_plan,
                    COALESCE(start_date, '') as start_date
                FROM students 
                ORDER BY full_name
            """, conn)
            conn.close()
            
            csv = df.to_csv(index=False)
            st.download_button(
                label="Download Students Data",
                data=csv,
                file_name=f"students_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
            

        
        if st.button("📥 Download Payments CSV"):
            conn = sqlite3.connect('chords_crm.db')
            df = pd.read_sql_query("SELECT * FROM payments", conn)
            conn.close()
            
            csv = df.to_csv(index=False)
            st.download_button(
                label="Download Payments Data",
                data=csv,
                file_name=f"payments_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
    
    st.divider()
    
    st.subheader("📤 Bulk Upload Students")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📋 Download Template**")
        if st.button("📥 Download Student Template Excel"):
            # Create Excel with dropdowns
            wb = Workbook()
            ws = wb.active
            ws.title = "Student Template"
            
            # Headers
            headers = ['full_name', 'age', 'mobile', 'email', 'date_of_birth', 'sex', 'instrument', 'class_plan', 'start_date']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add sample row with ddmmyyyy format
            ws.cell(row=2, column=1, value='John Doe')
            ws.cell(row=2, column=2, value=25)
            ws.cell(row=2, column=3, value='9876543210')
            ws.cell(row=2, column=4, value='john@example.com')
            ws.cell(row=2, column=5, value='15011998')
            ws.cell(row=2, column=6, value='Male')
            ws.cell(row=2, column=7, value='Piano')
            ws.cell(row=2, column=8, value='1 Month - 8')
            ws.cell(row=2, column=9, value='01012024')
            
            # Create separate sheets for dropdown lists
            sex_sheet = wb.create_sheet("SexOptions")
            sex_sheet['A1'] = "Male"
            sex_sheet['A2'] = "Female" 
            sex_sheet['A3'] = "Other"
            
            instrument_sheet = wb.create_sheet("InstrumentOptions")
            instruments = ["Piano", "Guitar", "Drums", "Violin", "Flute", "Keyboard", "Carnatic Vocals", "Hindustani Vocals", "Western Vocals"]
            for i, inst in enumerate(instruments, 1):
                instrument_sheet[f'A{i}'] = inst
            
            plan_sheet = wb.create_sheet("PlanOptions")
            plans = ["No Package", "1 Month - 8", "3 Month - 24", "6 Month - 48", "12 Month - 96"]
            for i, plan in enumerate(plans, 1):
                plan_sheet[f'A{i}'] = plan
            
            # Add data validation using sheet references
            sex_validation = DataValidation(type="list", formula1="SexOptions!$A$1:$A$3")
            sex_validation.add('F2:F1000')
            ws.add_data_validation(sex_validation)
            
            instrument_validation = DataValidation(type="list", formula1="InstrumentOptions!$A$1:$A$9")
            instrument_validation.add('G2:G1000')
            ws.add_data_validation(instrument_validation)
            
            plan_validation = DataValidation(type="list", formula1="PlanOptions!$A$1:$A$5")
            plan_validation.add('H2:H1000')
            ws.add_data_validation(plan_validation)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="Download Excel Template",
                data=excel_buffer.getvalue(),
                file_name="student_upload_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.info("💡 Excel template has dropdown lists in Sex, Instrument, and Class Plan columns. Click cells F2, G2, H2 to see dropdowns!")
        
        # Show dropdown options
        with st.expander("📝 View All Dropdown Options"):
            st.markdown("""
            **Instrument Options:**
            - Piano, Guitar, Drums, Violin, Flute, Keyboard
            - Carnatic Vocals, Hindustani Vocals, Western Vocals
            
            **Class Plan Options:**
            - No Package
            - 1 Month - 8
            - 3 Month - 24  
            - 6 Month - 48
            - 12 Month - 96
            
            **Gender Options:**
            - Male, Female, Other
            
            **Date Formats Accepted:**
            - YYYY-MM-DD (2024-01-15)
            - DD-MM-YYYY (15-01-2024)
            - DD/MM/YY (15/01/24)
            """)
    
    with col2:
        st.markdown("**📤 Upload Students**")
        uploaded_file = st.file_uploader("Choose file", type=["csv", "xlsx"])
        
        if uploaded_file is not None:
            try:
                # Read Excel or CSV
                if uploaded_file.name.endswith('.xlsx'):
                    df = pd.read_excel(uploaded_file)
                else:
                    df = pd.read_csv(uploaded_file)
                # Fill NaN values with empty strings
                df = df.fillna('')
                
                # Remove empty rows (where name is empty)
                df = df[df['full_name'].str.strip() != '']
                
                st.write(f"📊 Found {len(df)} students in uploaded file")
                if len(df) == 0:
                    st.error("No valid students found in file. Make sure 'full_name' column has data.")
                    return
                    
                st.dataframe(df)
                
                if st.button("✅ Upload Students", type="primary"):
                    success_count = 0
                    error_count = 0
                    
                    # Check for existing students to prevent duplicates
                    conn = sqlite3.connect('chords_crm.db')
                    cursor = conn.cursor()
                    
                    for _, row in df.iterrows():
                        try:
                            # Clean mobile number (remove decimal if present)
                            mobile_clean = str(row['mobile']).strip()
                            if '.' in mobile_clean:
                                mobile_clean = str(int(float(mobile_clean)))
                            
                            # Check if student already exists (by name and mobile)
                            cursor.execute('SELECT student_id FROM students WHERE full_name = ? AND mobile = ?', 
                                         (str(row['full_name']).strip(), mobile_clean))
                            existing = cursor.fetchone()
                            
                            if existing:
                                st.warning(f"⚠️ Student {row['full_name']} already exists, skipping...")
                                continue
                            
                            # Generate student ID
                            student_id = f"CMA{datetime.now().strftime('%Y%m%d%H%M%S')}{success_count:03d}"
                            
                            # Handle start date with multiple formats
                            start_date_str = str(row['start_date']).strip() if row['start_date'] else ''
                            if not start_date_str or start_date_str == 'nan':
                                start_date = datetime.now()
                            else:
                                # Handle Excel datetime format
                                if ' 00:00:00' in start_date_str:
                                    start_date_str = start_date_str.split(' ')[0]
                                try:
                                    # Try ddmmyyyy format first
                                    if len(start_date_str) == 8 and start_date_str.isdigit():
                                        start_date = datetime.strptime(start_date_str, '%d%m%Y')
                                    else:
                                        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                                except:
                                    try:
                                        start_date = datetime.strptime(start_date_str, '%d-%m-%Y')
                                    except:
                                        try:
                                            start_date = datetime.strptime(start_date_str, '%d/%m/%Y')
                                        except:
                                            start_date = datetime.now()
                            
                            # Parse date of birth with multiple formats
                            dob_str = str(row['date_of_birth']).strip() if row['date_of_birth'] else ''
                            if not dob_str or dob_str == 'nan':
                                dob_parsed = '2000-01-01'
                            else:
                                # Handle Excel datetime format
                                if ' 00:00:00' in dob_str:
                                    dob_str = dob_str.split(' ')[0]
                                try:
                                    # Try ddmmyyyy format first
                                    if len(dob_str) == 8 and dob_str.isdigit():
                                        dob_parsed = datetime.strptime(dob_str, '%d%m%Y').strftime('%Y-%m-%d')
                                    else:
                                        dob_parsed = datetime.strptime(dob_str, '%Y-%m-%d').strftime('%Y-%m-%d')
                                except:
                                    try:
                                        dob_parsed = datetime.strptime(dob_str, '%d-%m-%Y').strftime('%Y-%m-%d')
                                    except:
                                        try:
                                            dob_parsed = datetime.strptime(dob_str, '%d/%m/%Y').strftime('%Y-%m-%d')
                                        except:
                                            dob_parsed = '2000-01-01'
                            
                            # Calculate expiry date
                            package_days = {
                                "1 Month - 8": 30, "3 Month - 24": 90, 
                                "6 Month - 48": 180, "12 Month - 96": 365,
                                "No Package": 0
                            }
                            days = package_days.get(str(row['class_plan']).strip(), 30)
                            expiry_date = start_date + timedelta(days=days)
                            
                            # Validate and clean class plan
                            plan_str = str(row['class_plan']).strip() if row['class_plan'] else 'No Package'
                            if plan_str == 'nan' or not plan_str or plan_str.startswith('#'):
                                plan_str = 'No Package'
                            
                            # Validate against allowed plans
                            allowed_plans = ['No Package', '1 Month - 8', '3 Month - 24', '6 Month - 48', '12 Month - 96']
                            if plan_str not in allowed_plans:
                                st.warning(f"⚠️ Invalid plan '{plan_str}' for {row['full_name']}, using 'No Package'")
                                plan_str = 'No Package'
                            
                            total_classes = int(plan_str.split(' - ')[1]) if ' - ' in plan_str else 0
                            
                            # Skip comment/example rows
                            if str(row['full_name']).strip().startswith('#'):
                                continue
                                
                            # Handle age conversion safely
                            age_val = row['age'] if row['age'] and str(row['age']) != 'nan' else 18
                            try:
                                age_int = int(float(str(age_val)))
                            except:
                                age_int = 18
                            
                            cursor.execute('''
                                INSERT INTO students 
                                (student_id, full_name, age, mobile, email, date_of_birth, sex, 
                                 instrument, class_plan, total_classes, start_date, expiry_date)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                student_id, 
                                str(row['full_name']).strip() if row['full_name'] else 'Unknown', 
                                age_int,
                                mobile_clean, 
                                str(row['email']).strip() if row['email'] else '', 
                                dob_parsed, 
                                str(row['sex']).strip() if row['sex'] else 'Male', 
                                str(row['instrument']).strip() if row['instrument'] and not str(row['instrument']).startswith('#') else 'Piano', 
                                plan_str, total_classes,
                                start_date.strftime('%Y-%m-%d'), expiry_date.strftime('%Y-%m-%d')
                            ))
                            success_count += 1
                        except Exception as e:
                            error_count += 1
                            st.error(f"Error adding {row.get('full_name', 'Unknown')}: {str(e)}")
                    
                    conn.commit()
                    conn.close()
                    
                    if success_count > 0:
                        st.success(f"✅ Successfully added {success_count} students!")
                    if error_count > 0:
                        st.warning(f"⚠️ {error_count} students failed to upload")
                    
                    st.rerun()
            
            except Exception as e:
                st.error(f"Error reading file: {str(e)}")
    
    st.divider()
    
    st.subheader("⚠️ Important: Data Persistence")
    st.warning("""
    **Why students disappear after deployment:**
    - Streamlit Cloud resets files on each deployment
    - Only files in Git repository persist
    - New students added after deployment are lost unless committed to Git
    
    **Solution:**
    1. Export data regularly using buttons above
    2. For permanent storage, contact admin to commit database changes
    3. Consider upgrading to cloud database (PostgreSQL) for automatic persistence
    """)
    
    if st.button("🔄 Refresh Database Stats"):
        st.rerun()
    
    if st.button("← Back to Dashboard"):
        st.session_state.page = "dashboard"
        st.rerun()

# Add this to your main app.py navigation
def add_backup_to_navigation():
    """Add backup option to sidebar"""
    return st.sidebar.selectbox("Select Page", [
        "Add Student", "View Students", "Process Payment", 
        "Due Alerts", "Payment Overview", "Database Backup"
    ])